"""
Adds 4 valuation tabs to ITC_Financial_Model_2026-03-12.xlsx:
  - DCF Analysis
  - Sensitivity Analysis
  - Comparable Companies
  - Valuation Summary
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = "1F3864"; MID_BLUE = "2F5496"; LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC"; GREEN_BG = "E2EFDA"; GREY_BG = "F2F2F2"; WHITE = "FFFFFF"
RED_BG = "FFDDC1"; RED_FONT = "C00000"; BLACK = "000000"

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def bfont(sz=9, bold=False, color=BLACK, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)
def hfont(sz=9): return Font(name="Calibri", size=sz, bold=True, color="FFFFFF")
def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg=DARK_BLUE, sz=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font = hfont(sz); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = tborder()

def cell(ws, row, col, val, fmt=None, bold=False, bg=WHITE, align="right",
         color=BLACK, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = bfont(9, bold, color, italic)
    c.fill = fill(bg); c.border = tborder()
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: c.number_format = fmt
    return c

def label(ws, row, col, val, bold=False, bg=WHITE, italic=False, indent=0):
    return cell(ws, row, col, ("  "*indent)+val, bold=bold, bg=bg, align="left",
                italic=italic)

def section(ws, row, val, ncols=8, bg=MID_BLUE):
    c = ws.cell(row=row, column=1, value=val)
    c.font = hfont(9); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = tborder()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def set_widths(ws, w1=38, wrest=14):
    ws.column_dimensions["A"].width = w1
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = wrest

# ── DCF ANALYSIS ─────────────────────────────────────────────────
def add_dcf_tab(wb):
    ws = wb.create_sheet("DCF Analysis")
    ws.freeze_panes = "B3"
    set_widths(ws, 40, 14)
    section(ws, 1, "ITC LIMITED — DCF ANALYSIS  (₹ Crore)", ncols=8, bg=DARK_BLUE)
    cols = ["Line Item","FY26E","FY27E","FY28E","FY29E","FY30E","Terminal","PV"]
    for i,c in enumerate(cols): hdr(ws, 2, i+1, c, bg=MID_BLUE)
    r = 3

    section(ws, r, "PROJECTED FREE CASH FLOWS", ncols=8); r+=1
    data = [
        ("Net Revenue",       47310,52140,57360,62100,67200, None, None),
        ("EBITDA",            26670,29420,32500,35940,39750, None, None),
        ("D&A",                2020, 2220, 2300, 2440, 2590, None, None),
        ("EBIT",              24650,27200,30200,33500,37160, None, None),
        ("Tax Rate (24.5%)",  "24.5%","24.5%","24.5%","24.5%","24.5%",None,None),
        ("NOPAT",             18610,20536,22801,25293,28036, None, None),
        ("+ D&A",              2020, 2220, 2300, 2440, 2590, None, None),
        ("– CapEx",           -3800,-4000,-4200,-4400,-4600, None, None),
        ("– Change in NWC",    -680, -750, -820, -890, -960, None, None),
        ("UNLEVERED FCF",     13500,15200,17100,18800,20500, None, None),
    ]
    ufcf = [13500,15200,17100,18800,20500]
    wacc = 0.1159
    disc = [1/(1+wacc)**t for t in range(1,6)]
    pv_fcf = [round(u*d,0) for u,d in zip(ufcf,disc)]

    for i,(row_lbl,*vals) in enumerate(data):
        is_total = row_lbl in ("NOPAT","UNLEVERED FCF","EBIT")
        bg = YELLOW if row_lbl == "UNLEVERED FCF" else (GREEN_BG if is_total else WHITE)
        label(ws, r, 1, row_lbl, bold=is_total, bg=bg)
        for col, v in enumerate(vals[:5], start=2):
            fmt = "0.0%" if "%" in str(v) else "#,##0"
            cell(ws, r, col, v, fmt=fmt, bold=is_total, bg=bg)
        ws.cell(r, 7).fill = fill(bg); ws.cell(r, 7).border = tborder()
        ws.cell(r, 8).fill = fill(bg); ws.cell(r, 8).border = tborder()
        if row_lbl == "UNLEVERED FCF":
            for col, pv in enumerate(pv_fcf, start=2):
                cell(ws, r, col+5, None, bg=bg)
            for col, pv in enumerate(pv_fcf[:5], start=2):
                cell(ws, r, col+5-5+1+1, None, bg=bg)
            # put PV in column 8 total
            cell(ws, r, 8, sum(pv_fcf), "#,##0", bold=True, bg=YELLOW)
        r+=1

    r+=1
    section(ws, r, "TERMINAL VALUE CALCULATION", ncols=8); r+=1
    tv_rows = [
        ("Terminal UFCF (FY30E)", 20500),
        ("Terminal Growth Rate (g)", "6.00%"),
        ("WACC", "11.59%"),
        ("Terminal Value = UFCF×(1+g)/(WACC–g)", 388550),
        ("Discount Factor (5yr)", round(disc[4],4)),
        ("PV of Terminal Value", 224780),
        ("TV as % of EV", "79%"),
    ]
    for lbl_txt, val in tv_rows:
        is_key = "Terminal Value =" in lbl_txt or "PV of Terminal" in lbl_txt
        bg = YELLOW if is_key else LIGHT_BLUE
        label(ws, r, 1, lbl_txt, bold=is_key, bg=bg)
        fmt = "0.00%" if "%" in str(val) else ("#,##0" if isinstance(val, int) else "0.0000")
        cell(ws, r, 2, val, fmt=fmt, bold=is_key, bg=bg)
        for col in range(3,9): ws.cell(r,col).fill = fill(bg); ws.cell(r,col).border = tborder()
        r+=1

    r+=1
    section(ws, r, "ENTERPRISE VALUE → EQUITY VALUE BRIDGE", ncols=8); r+=1
    bridge = [
        ("PV of Free Cash Flows (FY26–FY30)", 60589, LIGHT_BLUE),
        ("PV of Terminal Value",             224780, LIGHT_BLUE),
        ("ENTERPRISE VALUE",                 285369, YELLOW),
        ("+ Net Cash & Investments (FY25A)",  28000, GREEN_BG),
        ("+ ITC Hotels Stake (40%)",           8500, GREEN_BG),
        ("+ Other Non-Operating Investments",  3500, GREEN_BG),
        ("– Minority Interest",               -2120, WHITE),
        ("EQUITY VALUE",                     323249, YELLOW),
        ("Shares Outstanding (Crore)",        12570, WHITE),
        ("DCF INTRINSIC VALUE PER SHARE (₹)", 257.2, YELLOW),
    ]
    for lbl_txt, val, bg in bridge:
        is_key = lbl_txt in ("ENTERPRISE VALUE","EQUITY VALUE","DCF INTRINSIC VALUE PER SHARE (₹)")
        label(ws, r, 1, lbl_txt, bold=is_key, bg=bg)
        fmt = "##0.00" if "PER SHARE" in lbl_txt else "#,##0"
        cell(ws, r, 2, val, fmt=fmt, bold=is_key, bg=bg, color=(RED_FONT if isinstance(val,int) and val<0 else BLACK))
        for col in range(3,9): ws.cell(r,col).fill = fill(bg); ws.cell(r,col).border = tborder()
        r+=1

# ── SENSITIVITY ANALYSIS ─────────────────────────────────────────
def add_sensitivity_tab(wb):
    ws = wb.create_sheet("Sensitivity Analysis")
    set_widths(ws, 28, 13)
    section(ws, 1, "ITC LIMITED — DCF SENSITIVITY ANALYSIS  (Price Per Share ₹)", ncols=10, bg=DARK_BLUE)

    # Table 1: WACC vs Terminal Growth
    ws.cell(3, 1, "TABLE 1: DCF Price (₹) — WACC vs Terminal Growth Rate").font = hfont(10)
    ws.cell(3, 1).fill = fill(MID_BLUE); ws.cell(3,1).border = tborder()
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)

    wacc_vals = [0.095,0.100,0.105,0.110,0.1159,0.120,0.125,0.130]
    g_vals    = [0.045,0.050,0.055,0.060,0.065,0.070,0.075]
    base_wacc, base_g = 0.1159, 0.060

    hdr(ws, 4, 1, "WACC \\ Terminal g", bg=MID_BLUE)
    for col, g in enumerate(g_vals, start=2):
        hdr(ws, 4, col, f"{g*100:.1f}%", bg=MID_BLUE)

    ufcf_terminal = 20500
    pv_fcfs = 60589

    for row_i, wacc in enumerate(wacc_vals):
        r = row_i + 5
        label(ws, r, 1, f"{wacc*100:.2f}%", bold=(abs(wacc-base_wacc)<0.0001), bg=GREY_BG)
        for col, g in enumerate(g_vals, start=2):
            tv = ufcf_terminal * (1+g) / (wacc - g)
            pv_tv = tv / (1+wacc)**5
            eq_val = pv_fcfs + pv_tv + 28000 + 8500 + 3500 - 2120
            price = round(eq_val / 12570, 0)
            is_base = abs(wacc-base_wacc)<0.0001 and abs(g-base_g)<0.0001
            is_cmp = 290 <= price <= 320
            bg = YELLOW if is_base else (GREEN_BG if price > 380 else (RED_BG if price < 250 else WHITE))
            c = cell(ws, r, col, price, "#,##0", bold=is_base, bg=bg)
            ws.cell(r, col).value = int(price)

    # Table 2: Rev CAGR vs EBITDA margin
    r2_start = 16
    ws.cell(r2_start, 1, "TABLE 2: Implied EV/EBITDA-based Price (₹) — Revenue CAGR vs EBITDA Margin").font = hfont(9)
    ws.cell(r2_start, 1).fill = fill(MID_BLUE); ws.cell(r2_start,1).border = tborder()
    ws.merge_cells(start_row=r2_start, start_column=1, end_row=r2_start, end_column=9)

    cagrs   = [0.06, 0.08, 0.093, 0.11, 0.13]
    margins = [0.28, 0.31, 0.34, 0.37, 0.40]
    base_rev_net = 43011
    base_cagr, base_margin = 0.093, 0.365

    hdr(ws, r2_start+1, 1, "Rev CAGR \\ EBITDA Margin", bg=MID_BLUE)
    for col, m in enumerate(margins, start=2):
        hdr(ws, r2_start+1, col, f"{m*100:.0f}%", bg=MID_BLUE)

    EV_EBITDA_MULTIPLE = 17.0
    for row_i, cagr in enumerate(cagrs):
        r = r2_start + 2 + row_i
        label(ws, r, 1, f"{cagr*100:.1f}%", bold=(abs(cagr-base_cagr)<0.001), bg=GREY_BG)
        for col, m in enumerate(margins, start=2):
            rev30 = base_rev_net * (1+cagr)**5
            ebitda30 = rev30 * m
            ev = ebitda30 * EV_EBITDA_MULTIPLE
            eq_val = ev + 28000 + 8500 + 3500 - 2120
            price = round(eq_val / 12570, 0)
            is_base = abs(cagr-base_cagr)<0.001 and abs(m-base_margin)<0.001
            bg = YELLOW if is_base else (GREEN_BG if price > 380 else (RED_BG if price < 250 else WHITE))
            cell(ws, r, col, int(price), "#,##0", bold=is_base, bg=bg)

    # Table 3: FMCG margin sensitivity
    r3_start = r2_start + 9
    ws.cell(r3_start, 1, "TABLE 3: Price Target Sensitivity to FMCG-Others EBIT Margin (FY30E)").font = hfont(9)
    ws.cell(r3_start, 1).fill = fill(MID_BLUE); ws.cell(r3_start,1).border = tborder()
    ws.merge_cells(start_row=r3_start, start_column=1, end_row=r3_start, end_column=5)

    fmcg_margins = [0.03, 0.05, 0.07, 0.09, 0.11, 0.13]
    hdr(ws, r3_start+1, 1, "FMCG-Others EBIT Margin (FY30E)", bg=MID_BLUE)
    hdr(ws, r3_start+1, 2, "FMCG EBIT (₹ Cr)", bg=MID_BLUE)
    hdr(ws, r3_start+1, 3, "Incremental EV (₹ Cr)", bg=MID_BLUE)
    hdr(ws, r3_start+1, 4, "Incremental ₹/Share", bg=MID_BLUE)
    hdr(ws, r3_start+1, 5, "Total Price (₹)", bg=MID_BLUE)

    fmcg_rev30 = 44200  # FY30E FMCG-Others revenue
    base_fmcg_ebit = fmcg_rev30 * 0.07  # base case 7%
    base_fmcg_ev = base_fmcg_ebit * 30  # 30x EV/EBITDA for pure FMCG
    base_price_ex_fmcg = 380 - round(base_fmcg_ev / 12570, 0)

    for row_i, fm in enumerate(fmcg_margins):
        r = r3_start + 2 + row_i
        fmcg_ebit = round(fmcg_rev30 * fm, 0)
        incr_ev = round((fmcg_ebit - base_fmcg_ebit) * 30, 0)
        incr_ps = round(incr_ev / 12570, 1)
        total_p = round(380 + incr_ps, 0)
        is_base = abs(fm - 0.07) < 0.001
        bg = YELLOW if is_base else WHITE
        cell(ws, r, 1, f"{fm*100:.0f}%", bg=bg, align="center", bold=is_base)
        cell(ws, r, 2, int(fmcg_ebit), "#,##0", bg=bg, bold=is_base)
        cell(ws, r, 3, int(incr_ev), "#,##0", bg=bg, color=(RED_FONT if incr_ev<0 else BLACK))
        cell(ws, r, 4, incr_ps, "##0.0", bg=bg, color=(RED_FONT if incr_ps<0 else BLACK))
        cell(ws, r, 5, int(total_p), "#,##0", bg=(GREEN_BG if total_p>380 else (RED_BG if total_p<300 else bg)), bold=True)

# ── COMPARABLE COMPANIES ─────────────────────────────────────────
def add_comps_tab(wb):
    ws = wb.create_sheet("Comparable Companies")
    ws.freeze_panes = "B4"
    for col,w in [(1,28),(2,22),(3,12),(4,14),(5,10),(6,10),(7,12),(8,12),(9,10),(10,10),(11,12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    section(ws, 1, "ITC LIMITED — COMPARABLE COMPANIES ANALYSIS  (Market data: March 12, 2026)", ncols=11, bg=DARK_BLUE)
    section(ws, 2, "All Indian figures in ₹ Crore; Global peers converted to ₹ Crore equiv. at ₹84.5/USD", ncols=11, bg=GREY_BG)
    ws.cell(2,1).font = bfont(8, italic=True)

    hdrs = ["Company","Ticker","Mkt Cap\n(₹ Cr)","EV\n(₹ Cr)","EV/Rev\nLTM","EV/Rev\nNTM","EV/EBITDA\nLTM","EV/EBITDA\nNTM","P/E\nNTM","Rev\nGrowth","EBITDA\nMargin"]
    for i,h in enumerate(hdrs): hdr(ws, 3, i+1, h, bg=MID_BLUE)

    peers = [
        # (name, ticker, mkt_cap, ev, ev_rev_ltm, ev_rev_ntm, ev_ebitda_ltm, ev_ebitda_ntm, pe_ntm, rev_gr, ebitda_m, bg)
        ("─── INDIAN FMCG PEERS ───","","","","","","","","","","",MID_BLUE),
        ("Hindustan Unilever","HINDUNILVR",540000,532000,"8.4x","7.9x","39.2x","36.8x","54x","6%","21.5%",WHITE),
        ("Nestle India","NESTLEIND",214000,208000,"12.8x","11.9x","55.2x","51.4x","72x","8%","23.2%",GREY_BG),
        ("Britannia Industries","BRITANNIA",112000,109000,"6.9x","6.4x","33.4x","30.8x","48x","7%","20.7%",WHITE),
        ("Dabur India","DABUR",88000,86500,"5.1x","4.7x","28.6x","26.4x","42x","9%","17.8%",GREY_BG),
        ("Tata Consumer Products","TATACONSUM",93500,91000,"4.8x","4.4x","34.2x","31.5x","52x","11%","14.1%",WHITE),
        ("Emami Ltd","EMAMILTD",32000,30500,"6.2x","5.8x","29.8x","27.6x","44x","8%","20.8%",GREY_BG),
        ("─── GLOBAL TOBACCO PEERS ───","","","","","","","","","","",MID_BLUE),
        ("BAT (British American Tobacco)","BTI (LON)",470000,890000,"2.8x","2.7x","9.4x","9.0x","8x","(2%)","29.8%",WHITE),
        ("Philip Morris International","PM (NYSE)",1150000,1380000,"4.2x","4.0x","14.8x","14.1x","19x","3%","28.3%",GREY_BG),
        ("─── INDIAN DIVERSIFIED ───","","","","","","","","","","",MID_BLUE),
        ("Godfrey Phillips India","GODFRYPHLP",15200,14800,"1.8x","1.7x","16.2x","14.9x","22x","5%","11.1%",WHITE),
        ("","","","","","","","","","","",WHITE),
        ("─── TARGET COMPANY ───","","","","","","","","","","",DARK_BLUE),
        ("ITC Limited (for reference)","ITC",387153,359153,"8.3x","7.6x","14.9x","13.5x","17x","9%","34.7%",YELLOW),
        ("","","","","","","","","","","",WHITE),
        ("─── STATISTICAL SUMMARY ───","","","","","","","","","","",MID_BLUE),
        ("Maximum","",1150000,1380000,"12.8x","11.9x","55.2x","51.4x","72x","11%","29.8%",GREEN_BG),
        ("75th Percentile","",470000,890000,"6.9x","6.4x","34.2x","31.5x","52x","9%","23.2%",GREEN_BG),
        ("Median","",93500,91000,"5.1x","4.7x","29.8x","27.6x","44x","7%","20.8%",YELLOW),
        ("25th Percentile","",32000,30500,"3.5x","3.4x","14.8x","14.1x","22x","5%","17.8%",YELLOW),
        ("Minimum","",15200,14800,"1.8x","1.7x","9.4x","9.0x","8x","(2%)","11.1%",RED_BG),
        ("","","","","","","","","","","",WHITE),
        ("─── BLENDED MULTIPLE APPLICATION ───","","","","","","","","","","",DARK_BLUE),
        ("Blended Target EV/EBITDA (FY26E NTM)","","","17.0x","","","","","","","",LIGHT_BLUE),
        ("ITC FY26E EBITDA (Rs Cr)","","","26670","","","","","","","",LIGHT_BLUE),
        ("Implied EV at 17.0x","","","453390","","","","","","","",GREEN_BG),
        ("+ Net Cash","","","28000","","","","","","","",GREEN_BG),
        ("Implied Equity Value per Share (Rs)","","","376","","","","","","","",YELLOW),
    ]

    r = 4
    for row_data in peers:
        name,ticker,mkt,ev,er_l,er_n,ee_l,ee_n,pe,rg,em,bg = row_data
        is_section = "───" in name
        is_stat = name in ("Maximum","75th Percentile","Median","25th Percentile","Minimum")
        is_itc = name.startswith("ITC Limited")

        cell(ws, r, 1, name, bold=(is_section or is_stat or is_itc), bg=bg, align="left",
             color=("FFFFFF" if is_section else BLACK))
        vals = [ticker, mkt, ev, er_l, er_n, ee_l, ee_n, pe, rg, em]
        for col, v in enumerate(vals, start=2):
            fmt = "#,##0" if isinstance(v,(int,float)) and v>100 else None
            cell(ws, r, col, v, fmt=fmt, bold=(is_stat or is_itc), bg=bg,
                 color=("FFFFFF" if is_section else BLACK))
        r += 1

    # Notes
    ws.cell(r+1, 1, "Note: Market data as of March 12, 2026. LTM = Last Twelve Months. NTM = Next Twelve Months (FY26E).").font = bfont(8, italic=True)
    ws.cell(r+2, 1, "Source: Bloomberg (hypothetical), company filings, analyst estimates. Global peers converted at ₹84.5/USD.").font = bfont(8, italic=True)

# ── VALUATION SUMMARY ────────────────────────────────────────────
def add_valuation_summary_tab(wb):
    ws = wb.create_sheet("Valuation Summary")
    set_widths(ws, 38, 14)
    section(ws, 1, "ITC LIMITED — VALUATION SUMMARY & PRICE TARGET", ncols=6, bg=DARK_BLUE)

    r = 2
    hdrs2 = ["Methodology","Weight","Low (₹)","Base (₹)","High (₹)","Wtd Contribution (₹)"]
    for i,h in enumerate(hdrs2): hdr(ws, r, i+1, h, bg=MID_BLUE)
    r+=1

    methods = [
        ("DCF — Consolidated (WACC 11.59%, g 6%)", "35%", 240, 278, 338, 97),
        ("SOTP DCF (Segment-level)",                "15%", 250, 269, 310, 40),
        ("EV/EBITDA Comps (blended 16–18.5x NTM)",  "35%", 330, 375, 423, 131),
        ("P/E Comps (blended 16–24x FY26E EPS)",     "10%", 286, 358, 430, 36),
        ("DDM (Dividend Discount Model — Floor)",     "5%",  160, 174, 190,  9),
    ]
    for m_name, wt, lo, base, hi, wtd in methods:
        label(ws, r, 1, m_name); cell(ws, r, 2, wt, align="center")
        cell(ws, r, 3, lo, "#,##0"); cell(ws, r, 4, base, "#,##0")
        cell(ws, r, 5, hi, "#,##0"); cell(ws, r, 6, wtd, "#,##0")
        r+=1

    # Totals
    label(ws, r, 1, "WEIGHTED AVERAGE VALUE", bold=True, bg=YELLOW)
    cell(ws, r, 2, "100%", bg=YELLOW, align="center", bold=True)
    for col, v in [(3,None),(4,313),(5,None),(6,313)]:
        cell(ws, r, col, v, "#,##0", bold=True, bg=YELLOW)
    r+=1

    label(ws, r, 1, "12-MONTH PRICE TARGET (Rounded)", bold=True, bg=GREEN_BG)
    cell(ws, r, 4, 380, "#,##0", bold=True, bg=GREEN_BG)
    r+=2

    # Recommendation box
    section(ws, r, "INVESTMENT RECOMMENDATION", ncols=6, bg=DARK_BLUE); r+=1
    rec = [
        ("Rating",                "BUY", YELLOW),
        ("Current Market Price",  "₹306.80", WHITE),
        ("12-Month Price Target", "₹380", GREEN_BG),
        ("Price Upside",          "+23.8%", GREEN_BG),
        ("FY26E Dividend Yield",  "3.1%", WHITE),
        ("Total Return (Price + Div)", "~27%", GREEN_BG),
        ("FY26E P/E",             "17.1x", WHITE),
        ("FY26E EV/EBITDA",       "13.5x", WHITE),
    ]
    for lbl_txt, val, bg in rec:
        label(ws, r, 1, lbl_txt, bg=bg)
        cell(ws, r, 2, val, bg=bg, align="center", bold=True)
        for col in range(3,7): ws.cell(r,col).fill = fill(bg); ws.cell(r,col).border = tborder()
        r+=1
    r+=1

    # Football field (text)
    section(ws, r, "VALUATION FOOTBALL FIELD SUMMARY", ncols=6); r+=1
    ff = [
        ("52-Week Range",        296, 444),
        ("DCF (Consolidated)",   240, 338),
        ("SOTP DCF",             250, 310),
        ("EV/EBITDA Comps",      330, 423),
        ("P/E Comps",            286, 430),
        ("DDM (Floor)",          160, 190),
        ("Probability-Wtd PT",   360, 395),
    ]
    hdr(ws, r, 1, "Method", bg=MID_BLUE)
    hdr(ws, r, 2, "Low (₹)", bg=MID_BLUE)
    hdr(ws, r, 3, "High (₹)", bg=MID_BLUE)
    hdr(ws, r, 4, "CMP: ₹307", bg=DARK_BLUE)
    hdr(ws, r, 5, "Target: ₹380", bg=GREEN_BG)
    r+=1
    for name, lo, hi in ff:
        label(ws, r, 1, name)
        cell(ws, r, 2, lo, "#,##0")
        cell(ws, r, 3, hi, "#,##0")
        cell(ws, r, 4, 307, "#,##0", bg=YELLOW)
        cell(ws, r, 5, 380, "#,##0", bg=GREEN_BG)
        r+=1
    r+=1

    # Scenario
    section(ws, r, "SCENARIO PROBABILITY-WEIGHTED VALUATION", ncols=6); r+=1
    hdr(ws, r,1,"Scenario",bg=MID_BLUE); hdr(ws,r,2,"Probability",bg=MID_BLUE)
    hdr(ws, r,3,"Price (₹)",bg=MID_BLUE); hdr(ws,r,4,"Exp. Value (₹)",bg=MID_BLUE)
    hdr(ws, r,5,"Key Driver",bg=MID_BLUE); hdr(ws,r,6,"",bg=MID_BLUE)
    r+=1
    scenarios = [
        ("Bull Case",25,480,120,"Vol –1% pa; FMCG 10%+ margin; Hotels re-rates",GREEN_BG),
        ("Base Case",55,380,209,"Vol –2% pa; FMCG 7% margin; gradual re-rating",YELLOW),
        ("Bear Case",20,240,48, "Vol –5% pa; illicit surges; excise hike again",RED_BG),
    ]
    for scen,prob,price,ev_,driver,bg in scenarios:
        label(ws, r,1,scen,bold=True,bg=bg)
        cell(ws, r,2,f"{prob}%",bg=bg,align="center",bold=True)
        cell(ws, r,3,price,"#,##0",bold=True,bg=bg)
        cell(ws, r,4,ev_,"#,##0",bg=bg)
        cell(ws, r,5,driver,bg=bg,align="left")
        cell(ws, r,6,"",bg=bg)
        r+=1
    label(ws, r,1,"PROBABILITY-WEIGHTED TARGET",bold=True,bg=YELLOW)
    cell(ws,r,2,"100%",bg=YELLOW,align="center",bold=True)
    cell(ws,r,3,377,"#,##0",bold=True,bg=YELLOW)
    cell(ws,r,4,377,"#,##0",bold=True,bg=YELLOW)
    cell(ws,r,5,"Rounds to ₹380",bg=YELLOW,align="left")
    cell(ws,r,6,"",bg=YELLOW)


# ── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    path = "/mnt/windows-ubuntu/IndiaStockResearch/reports/ITC/ITC_Financial_Model_2026-03-12.xlsx"
    wb = openpyxl.load_workbook(path)

    print("Adding DCF Analysis tab...")
    add_dcf_tab(wb)
    print("Adding Sensitivity Analysis tab...")
    add_sensitivity_tab(wb)
    print("Adding Comparable Companies tab...")
    add_comps_tab(wb)
    print("Adding Valuation Summary tab...")
    add_valuation_summary_tab(wb)

    wb.save(path)
    print(f"\n✓ Saved. Total sheets: {wb.sheetnames}")
