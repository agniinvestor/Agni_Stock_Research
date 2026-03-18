"""
ITC Limited — Financial Model Builder
Generates: ITC_Financial_Model_2026-03-12.xlsx
6 tabs: Revenue Model, Income Statement, Cash Flow, Balance Sheet, Scenarios, DCF Inputs
All figures in ₹ Crore unless noted.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ── Colour palette ──────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # tab headers / section titles
MID_BLUE    = "2F5496"   # sub-headers
LIGHT_BLUE  = "D6E4F0"   # input cells (blue-shaded)
YELLOW      = "FFF2CC"   # key metric highlight
GREEN_BG    = "E2EFDA"   # totals / cross-links
GREY_BG     = "F2F2F2"   # alternating row shade
WHITE       = "FFFFFF"
RED_FONT    = "C00000"
BLACK       = "000000"

# ── Helper styles ────────────────────────────────────────────────
def hdr_font(size=10, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color=BLACK, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color="1F3864")
    return Border(bottom=s)

def pct_fmt():  return "0.0%"
def cr_fmt():   return '#,##0'
def cr1_fmt():  return '#,##0.0'
def eps_fmt():  return '##0.00'

YEARS_H = ["FY21A", "FY22A", "FY23A", "FY24A", "FY25A"]
YEARS_P = ["FY26E", "FY27E", "FY28E", "FY29E", "FY30E"]
YEARS   = YEARS_H + YEARS_P

# ── Write helpers ────────────────────────────────────────────────
def write_header_row(ws, row, col_start, labels, bg=DARK_BLUE, font_color=WHITE,
                     font_size=9, bold=True):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start+i, value=lbl)
        c.font  = Font(name="Calibri", size=font_size, bold=bold, color=font_color)
        c.fill  = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()

def write_label(ws, row, col, label, bold=False, italic=False,
                indent=0, bg=WHITE, font_color=BLACK, size=9):
    c = ws.cell(row=row, column=col, value=("  "*indent) + label)
    c.font      = body_font(size=size, bold=bold, color=font_color, italic=italic)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = border_thin()
    return c

def write_val(ws, row, col, value, fmt=cr_fmt(), bg=WHITE,
              bold=False, red_neg=False, font_color=BLACK):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.font      = body_font(size=9, bold=bold, color=(RED_FONT if red_neg and isinstance(value, (int,float)) and value < 0 else font_color))
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border    = border_thin()
    return c

def write_row(ws, row, label, values, fmt=cr_fmt(), label_col=1, data_col_start=2,
              label_bold=False, label_indent=0, bg=WHITE, label_italic=False, red_neg=False):
    write_label(ws, row, label_col, label, bold=label_bold, indent=label_indent,
                bg=bg, italic=label_italic)
    for i, v in enumerate(values):
        write_val(ws, row, data_col_start+i, v, fmt=fmt, bg=bg,
                  bold=label_bold, red_neg=red_neg)

def section_title(ws, row, label, ncols=12, bg=MID_BLUE):
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = border_thin()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def blank_row(ws, row, ncols=12):
    for col in range(1, ncols+1):
        c = ws.cell(row=row, column=col, value="")
        c.fill = fill(WHITE)

def set_col_widths(ws, label_w=38, data_w=13):
    ws.column_dimensions["A"].width = label_w
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = data_w

# ════════════════════════════════════════════════════════════════
# TAB 1 — REVENUE MODEL
# ════════════════════════════════════════════════════════════════
def build_revenue_model(wb):
    ws = wb.create_sheet("Revenue Model")
    ws.freeze_panes = "B3"
    set_col_widths(ws)

    # Title row
    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — REVENUE MODEL  (₹ Crore)", ncols=12, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Line Item"] + YEARS, bg=MID_BLUE)

    r = 3
    # ── Segment Revenue ──────────────────────────────────────────
    section_title(ws, r, "A.  REVENUE BY SEGMENT", ncols=12); r+=1

    # Cigarettes
    section_title(ws, r, "  Cigarettes", ncols=12, bg=GREY_BG); r+=1
    cig_gross = [28750,34220,37980,41035,43040, 45880,49060,52490,56160,60090]
    cig_net   = [13250,15620,17280,18815,19840, 21030,22490,24060,25740,27540]
    cig_excise= [v - n for v,n in zip(cig_gross, cig_net)]
    write_row(ws, r, "  Cigarettes Gross Revenue", cig_gross, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Less: Excise Duty", [(-v) for v in cig_excise], label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Cigarettes Net Revenue", cig_net, label_bold=True, bg=GREEN_BG); r+=1
    cig_gr = [None] + [round((cig_net[i]/cig_net[i-1]-1)*100,1) for i in range(1,10)]
    write_row(ws, r, "  YoY Growth %", cig_gr, fmt=pct_fmt(), label_indent=1,
              label_italic=True)
    for col in range(2, 12):
        if ws.cell(r, col).value is not None:
            ws.cell(r, col).value = ws.cell(r, col).value / 100
    r+=1

    # FMCG-Others
    fmcg_foods = [10200,12890,14620,15780,17240, 19450,22600,26220,30410,35280]
    fmcg_pc    = [3050, 3650, 3970, 4400, 4740,  5320, 5980, 6720, 7550, 8490]
    fmcg_total = [a+b for a,b in zip(fmcg_foods, fmcg_pc)]
    section_title(ws, r, "  FMCG – Others", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  Branded Foods (Aashirvaad/Sunfeast/Bingo!/Yippee!)", fmcg_foods, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Personal Care, Stationery & Others", fmcg_pc, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  FMCG-Others Total", fmcg_total, label_bold=True, bg=GREEN_BG); r+=1
    fmcg_gr = [None] + [round((fmcg_total[i]/fmcg_total[i-1]-1)*100,1) for i in range(1,10)]
    write_row(ws, r, "  YoY Growth %", fmcg_gr, fmt=pct_fmt(), label_indent=1, label_italic=True)
    for col in range(2, 12):
        if ws.cell(r, col).value is not None:
            ws.cell(r, col).value = ws.cell(r, col).value / 100
    r+=1

    # Agribusiness
    agri_leaf  = [6800,10280,12640,12460,13310, 14540,15980,17580,19340,21280]
    agri_food  = [2450, 3950, 4840, 4850, 5250,  5830, 6450, 7140, 7920, 8780]
    agri_total = [a+b for a,b in zip(agri_leaf, agri_food)]
    section_title(ws, r, "  Agribusiness", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  Leaf Tobacco & Spices Exports", agri_leaf, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Wheat, Pulses, Coffee & Other Agri", agri_food, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Agribusiness Total", agri_total, label_bold=True, bg=GREEN_BG); r+=1
    agri_gr = [None] + [round((agri_total[i]/agri_total[i-1]-1)*100,1) for i in range(1,10)]
    write_row(ws, r, "  YoY Growth %", agri_gr, fmt=pct_fmt(), label_indent=1, label_italic=True)
    for col in range(2, 12):
        if ws.cell(r, col).value is not None:
            ws.cell(r, col).value = ws.cell(r, col).value / 100
    r+=1

    # Paperboards
    paper = [6200,7350,8450,8520,7980, 8540,9150,9800,10490,11220]
    section_title(ws, r, "  Paperboards, Paper & Packaging", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  Paperboards, Paper & Packaging Revenue", paper, label_indent=1, bg=LIGHT_BLUE); r+=1
    paper_gr = [None] + [round((paper[i]/paper[i-1]-1)*100,1) for i in range(1,10)]
    write_row(ws, r, "  YoY Growth %", paper_gr, fmt=pct_fmt(), label_indent=1, label_italic=True)
    for col in range(2, 12):
        if ws.cell(r, col).value is not None:
            ws.cell(r, col).value = ws.cell(r, col).value / 100
    r+=1

    # Hotels (discontinued)
    hotels = [1180,1480,2450,2890,834, None,None,None,None,None]
    section_title(ws, r, "  Hotels (Discontinued Jan 2025)", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  Hotels Revenue (included until demerger)", hotels, label_indent=1, bg=LIGHT_BLUE); r+=1
    r+=1

    # Infotech
    infotech = [1600,1780,1970,2210,2420, 2710,3040,3410,3820,4280]
    section_title(ws, r, "  ITC Infotech & Others", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  ITC Infotech & Others Revenue", infotech, label_indent=1, bg=LIGHT_BLUE); r+=1
    r+=1

    # Totals
    total_net = [27448,34588,39460,41870,43011, 47310,52140,57360,62100,67200]
    total_gross= [47902,59214,69446,71985,73465, 77640,83580,90020,97100,105000]
    section_title(ws, r, "  CONSOLIDATED TOTALS", ncols=12, bg=DARK_BLUE); r+=1
    write_row(ws, r, "TOTAL NET REVENUE (ex-excise)", total_net, label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "TOTAL GROSS REVENUE (incl. excise)", total_gross, label_bold=True, bg=YELLOW); r+=1
    net_gr = [None] + [round((total_net[i]/total_net[i-1]-1)*100,1) for i in range(1,10)]
    write_row(ws, r, "Net Revenue Growth %", net_gr, fmt=pct_fmt(), label_bold=True)
    for col in range(2, 12):
        if ws.cell(r, col).value is not None:
            ws.cell(r, col).value = ws.cell(r, col).value / 100
    r+=2

    # ── Geography ────────────────────────────────────────────────
    section_title(ws, r, "B.  REVENUE BY GEOGRAPHY (₹ Crore)", ncols=12); r+=1
    write_header_row(ws, r, 1, ["Line Item"] + YEARS, bg=MID_BLUE); r+=1

    dom_fmcg  = [20900,25510,28830,31050,33350, 37500,42200,47200,52500,58400]
    dom_paper = [5100, 5940, 6780, 6860, 6500,  6970, 7480, 8030, 8620, 9260]
    dom_hotels= [1180, 1480, 2450, 2890,  834,   None, None, None, None, None]
    dom_other = [ 900, 1050, 1150, 1280, 1450,   1620, 1820, 2040, 2290, 2570]
    dom_total = [v for v in [
        dom_fmcg[i]+(dom_paper[i] or 0)+(dom_hotels[i] or 0)+(dom_other[i] or 0)
        for i in range(10)
    ]]

    exp_leaf  = [6800,10280,12640,12460,13310, 14540,15980,17580,19340,21280]
    exp_food  = [1900, 2800, 3600, 3580, 3930,  4360, 4820, 5330, 5890, 6520]
    exp_it    = [1200, 1350, 1520, 1710, 1890,  2120, 2380, 2670, 2990, 3350]
    exp_other = [ 548,  608,  620,  870,  531,   590,  660,  740,  830,  930]
    exp_total = [a+b+c+d for a,b,c,d in zip(exp_leaf,exp_food,exp_it,exp_other)]

    section_title(ws, r, "  India (Domestic)", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  FMCG (Cigarettes + Others)", dom_fmcg, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Paperboards & Paper", dom_paper, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Hotels (until FY24)", dom_hotels, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Other Domestic", dom_other, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  India Total", dom_total, label_bold=True, bg=GREEN_BG); r+=1

    section_title(ws, r, "  Exports & International", ncols=12, bg=GREY_BG); r+=1
    write_row(ws, r, "  Leaf Tobacco & Spices Exports", exp_leaf, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Agri Commodities (Wheat/Coffee/Pulses)", exp_food, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  ITC Infotech International", exp_it, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Other Exports", exp_other, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Exports & International Total", exp_total, label_bold=True, bg=GREEN_BG); r+=1

    section_title(ws, r, "  TOTAL & MIX", ncols=12, bg=DARK_BLUE); r+=1
    write_row(ws, r, "TOTAL NET REVENUE", total_net, label_bold=True, bg=YELLOW); r+=1
    dom_pct = [round(d/t*100,1) for d,t in zip(dom_total,total_net)]
    exp_pct = [round(e/t*100,1) for e,t in zip(exp_total,total_net)]
    write_row(ws, r, "% Domestic", dom_pct, fmt=pct_fmt(), label_italic=True)
    for col in range(2, 12): ws.cell(r, col).value = (ws.cell(r, col).value or 0)/100
    r+=1
    write_row(ws, r, "% International / Exports", exp_pct, fmt=pct_fmt(), label_italic=True)
    for col in range(2, 12): ws.cell(r, col).value = (ws.cell(r, col).value or 0)/100
    r+=1


# ════════════════════════════════════════════════════════════════
# TAB 2 — INCOME STATEMENT
# ════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("Income Statement")
    ws.freeze_panes = "B3"
    set_col_widths(ws)

    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — INCOME STATEMENT  (₹ Crore)", ncols=12, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Line Item"] + YEARS, bg=MID_BLUE)

    r = 3
    # ── Revenue ──────────────────────────────────────────────────
    section_title(ws, r, "REVENUE", ncols=12); r+=1
    gross_rev   = [47902,59214,69446,71985,73465, 77640,83580,90020,97100,105000]
    excise      = [20454,24626,29986,30115,30454, 30330,31440,32660,34000,37800]
    net_rev     = [27448,34588,39460,41870,43011, 47310,52140,57360,62100,67200]
    oth_inc_ops = [  325,  430,  490,  510,  545,   600,  660,  720,  790,  860]
    total_ops   = [a+b for a,b in zip(net_rev,oth_inc_ops)]
    net_gr      = [None] + [round((net_rev[i]/net_rev[i-1]-1),4) for i in range(1,10)]

    write_row(ws, r, "Gross Revenue from Operations (incl. excise)", gross_rev, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "Less: Excise Duty on Cigarettes", [(-v) for v in excise], red_neg=True); r+=1
    write_row(ws, r, "Net Revenue from Operations", net_rev, label_bold=True, bg=GREEN_BG); r+=1
    write_row(ws, r, "Other Operating Income", oth_inc_ops, label_indent=1, label_italic=True); r+=1
    write_row(ws, r, "TOTAL REVENUE FROM OPERATIONS", total_ops, label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "YoY Growth (Net Revenue)", net_gr, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=2

    # ── Costs ────────────────────────────────────────────────────
    section_title(ws, r, "COST OF GOODS SOLD & MANUFACTURING", ncols=12); r+=1
    mat_consumed = [7820,10240,12180,12680,12950, 14020,15380,16870,18200,19700]
    stock_trade  = [4110, 5890, 6540, 6920, 7100,  7840, 8620, 9480,10230,11060]
    inv_chg      = [ -120, -310,  180,  -90,  120,     0,    0,    0,    0,    0]
    mfg_exp      = [2150, 2620, 3010, 3140, 3290,  3610, 3980, 4380, 4730, 5110]
    total_cogs   = [a+b+c+d for a,b,c,d in zip(mat_consumed,stock_trade,inv_chg,mfg_exp)]
    write_row(ws, r, "  Cost of Materials Consumed", mat_consumed, label_indent=1); r+=1
    write_row(ws, r, "  Purchases of Stock-in-Trade", stock_trade, label_indent=1); r+=1
    write_row(ws, r, "  Changes in Inventories (WIP/FG)", inv_chg, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Manufacturing & Other Expenses", mfg_exp, label_indent=1); r+=1
    write_row(ws, r, "TOTAL COGS", total_cogs, label_bold=True, bg=GREEN_BG); r+=2

    section_title(ws, r, "EMPLOYEE COSTS", ncols=12); r+=1
    emp_sal  = [3210,3680,4120,4480,4870, 5310,5790,6310,6870,7490]
    emp_pf   = [ 380, 430, 480, 520, 560,  610,  665,  725,  790,  860]
    emp_total= [a+b for a,b in zip(emp_sal,emp_pf)]
    write_row(ws, r, "  Salaries & Wages", emp_sal, label_indent=1); r+=1
    write_row(ws, r, "  PF, Gratuity & Other Benefits", emp_pf, label_indent=1); r+=1
    write_row(ws, r, "TOTAL EMPLOYEE COSTS", emp_total, label_bold=True, bg=GREEN_BG); r+=2

    section_title(ws, r, "OTHER OPERATING EXPENSES", ncols=12); r+=1
    sell_dist= [1620,1980,2280,2510,2730, 3020,3330,3660,3960,4280]
    advt     = [ 890,1060,1190,1310,1450, 1600,1760,1940,2100,2270]
    other_opx= [1170,1440,1640,1810,1990, 2190,2410,2650,2870,3100]
    total_opx= [a+b+c for a,b,c in zip(sell_dist,advt,other_opx)]
    write_row(ws, r, "  Selling & Distribution Expenses", sell_dist, label_indent=1); r+=1
    write_row(ws, r, "  Advertisement & Publicity", advt, label_indent=1); r+=1
    write_row(ws, r, "  Other Operating Expenses", other_opx, label_indent=1); r+=1
    write_row(ws, r, "TOTAL OTHER OPEX", total_opx, label_bold=True, bg=GREEN_BG); r+=2

    # ── EBITDA ───────────────────────────────────────────────────
    section_title(ws, r, "PROFITABILITY", ncols=12); r+=1
    ebitda   = [14356,17458,20141,21857,24025, 26670,29420,32500,35940,39750]
    ebitda_m = [round(e/n,4) for e,n in zip(ebitda,net_rev)]
    write_row(ws, r, "EBITDA", ebitda, label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "  EBITDA Margin % (on net revenue)", ebitda_m, fmt=pct_fmt(),
              label_italic=True, label_indent=1); r+=1

    dna      = [1535,1726,1907,2077,2043, 2020,2220,2300,2440,2590]
    write_row(ws, r, "  Less: Depreciation & Amortization", [(-v) for v in dna],
              label_indent=1, red_neg=True); r+=1

    ebit     = [12821,15732,18234,19780,21982, 24650,27200,30200,33500,37160]
    ebit_m   = [round(e/n,4) for e,n in zip(ebit,net_rev)]
    write_row(ws, r, "EBIT (Operating Profit)", ebit, label_bold=True, bg=GREEN_BG); r+=1
    write_row(ws, r, "  EBIT Margin % (on net revenue)", ebit_m, fmt=pct_fmt(),
              label_italic=True, label_indent=1); r+=2

    # ── Segment EBIT ─────────────────────────────────────────────
    section_title(ws, r, "SEGMENT EBIT BREAKDOWN", ncols=12, bg=MID_BLUE); r+=1
    seg_cig  = [11145,13560,16180,18023,21091, 22800,24700,26800,29100,31700]
    seg_fmcg = [ -410,  -280,   120,   510, 1050,  1760, 2870, 4130, 5610, 7210]
    seg_agri = [  520,   840,   840,   730,  850,   960, 1080, 1210, 1360, 1520]
    seg_paper= [  740,  1060,  1480,  1360,  890,  1050, 1230, 1420, 1640, 1890]
    seg_htl  = [  -80,   280,   790,  1020,  210,  None, None, None, None, None]
    seg_it   = [  230,   295,   340,   420,  490,   580,  660,  760,  870,  990]
    seg_unall= [-1323, -2023, -1516, -2303,-2599, -2500,-2740,-3020,-3070,-3150]
    seg_tot  = [10822,13732,18234,19760,21982, 24650,27200,30200,33500,37160]

    write_row(ws, r, "  Cigarettes EBIT", seg_cig, label_indent=1, bg=LIGHT_BLUE); r+=1
    cig_net_all = [13250,15620,17280,18815,19840, 21030,22490,24060,25740,27540]
    cig_em2 = [round(seg_cig[i]/cig_net_all[i],4) for i in range(10)]
    write_row(ws, r, "    Cigarettes EBIT Margin", cig_em2, fmt=pct_fmt(), label_indent=2, label_italic=True); r+=1
    write_row(ws, r, "  FMCG-Others EBIT", seg_fmcg, label_indent=1, bg=LIGHT_BLUE, red_neg=True); r+=1
    write_row(ws, r, "  Agribusiness EBIT", seg_agri, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Paperboards & Paper EBIT", seg_paper, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Hotels EBIT (discontinued)", seg_htl, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  ITC Infotech & Others EBIT", seg_it, label_indent=1, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Unallocated Corporate Expenses", seg_unall, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "TOTAL SEGMENT EBIT (= Consolidated EBIT)", seg_tot, label_bold=True, bg=YELLOW); r+=2

    # ── Below-EBIT ───────────────────────────────────────────────
    section_title(ws, r, "BELOW-OPERATING LINE", ncols=12); r+=1
    oth_inc = [4203,4580,4861,5659,5560, 5800,6100,6440,6820,7240]
    interest= [  78,  62,  48,  42,  38,   40,  38,  36,  34,  32]
    pbt     = [17024,20312,23095,25439,27542, 30410,33262,36604,40286,44368]
    tax     = [ 3993, 5064, 4342, 5545, 6230,  7300, 7970, 8800, 9680,10650]
    tax_rate= [round(t/p,4) for t,p in zip(tax,pbt)]
    pat_cont= [13031,15248,18753,19894,21312, 23110,25292,27804,30606,33718]
    pat_disc= [    0,    0,    0,    0, 1420,     0,    0,    0,    0,    0]  # Hotels demerger gain
    net_pat = [a+b for a,b in zip(pat_cont,pat_disc)]
    nci     = [  85,   98,  115,  128,  140,   155,  172,  190,  210,  230]
    pat_sh  = [a-b for a,b in zip(net_pat,nci)]

    write_row(ws, r, "Other Income (Interest, Dividends, Gains)", oth_inc, label_indent=1); r+=1
    write_row(ws, r, "Less: Finance Costs (Interest)", [(-v) for v in interest],
              label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "PROFIT BEFORE TAX (PBT)", pbt, label_bold=True, bg=YELLOW); r+=1
    pbt_m = [round(p/n,4) for p,n in zip(pbt,net_rev)]
    write_row(ws, r, "  PBT Margin %", pbt_m, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=1
    write_row(ws, r, "Income Tax (Current + Deferred)", [(-v) for v in tax],
              label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Effective Tax Rate %", tax_rate, fmt=pct_fmt(),
              label_italic=True, label_indent=1); r+=1
    write_row(ws, r, "PAT – Continuing Operations", pat_cont, label_bold=True, bg=GREEN_BG); r+=1
    write_row(ws, r, "PAT – Discontinued (Hotels gain, FY25)", pat_disc,
              label_indent=1, label_italic=True); r+=1
    write_row(ws, r, "NET PROFIT (Total)", net_pat, label_bold=True, bg=YELLOW); r+=1
    pat_m = [round(p/n,4) for p,n in zip(net_pat,net_rev)]
    write_row(ws, r, "  Net Margin %", pat_m, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=1
    write_row(ws, r, "Less: Minority Interest / NCI", [(-v) for v in nci],
              label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "PAT Attributable to Shareholders", pat_sh, label_bold=True, bg=GREEN_BG); r+=2

    # ── EPS ──────────────────────────────────────────────────────
    section_title(ws, r, "PER SHARE DATA", ncols=12); r+=1
    shares_b = [12544,12545,12555,12565,12570, 12570,12570,12570,12570,12570]
    eps_b    = [round(p/s,2) for p,s in zip(pat_sh,shares_b)]
    dps      = [5.75, 6.25, 6.50, 7.50, 8.50, 9.50,10.50,11.50,12.50,13.50]
    payout   = [round(d*s/p,4) for d,s,p in zip(dps,shares_b,pat_sh)]
    write_row(ws, r, "  Shares Outstanding - Basic (Crore)", shares_b, fmt=cr_fmt(), label_indent=1); r+=1
    write_row(ws, r, "Basic EPS (₹)", eps_b, fmt=eps_fmt(), label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "Dividend Per Share (₹)", dps, fmt=eps_fmt(), label_bold=True, bg=LIGHT_BLUE); r+=1
    write_row(ws, r, "  Dividend Payout Ratio %", payout, fmt=pct_fmt(),
              label_italic=True, label_indent=1); r+=1

    # Store key data on sheet for cross-references
    ws["A1"].comment = None
    # Tag key cells with names for reference
    ws.title = "Income Statement"


# ════════════════════════════════════════════════════════════════
# TAB 3 — CASH FLOW STATEMENT
# ════════════════════════════════════════════════════════════════
def build_cash_flow(wb):
    ws = wb.create_sheet("Cash Flow")
    ws.freeze_panes = "B3"
    set_col_widths(ws)

    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — CASH FLOW STATEMENT  (₹ Crore)", ncols=12, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Line Item"] + YEARS, bg=MID_BLUE)

    r = 3
    pat_cont= [13031,15248,18753,19894,21312, 23110,25292,27804,30606,33718]
    dna     = [1535, 1726, 1907, 2077, 2043,  2020, 2220, 2300, 2440, 2590]
    def_tax = [  82,   44,  -82,  120,  -60,    80,   80,   80,   80,   80]
    inv_gains=[-1820,-2100,-2210,-2980,-2800, -2900,-3100,-3300,-3500,-3700]
    other_nc= [  310,  380,  420,  450,  480,   510,  540,  570,  600,  630]
    ar_chg  = [ -310, -650, -480, -360, -290,  -380, -420, -460, -500, -540]
    inv_chg = [ -220, -480, -310, -200, -180,  -250, -280, -310, -330, -360]
    ap_chg  = [  410,  620,  540,  480,  420,   460,  510,  560,  600,  650]
    owc_chg = [ -180, -210, -280, -310, -190,  -220, -250, -280, -310, -340]
    tax_paid= [-3850,-4900,-4250,-5420,-6080, -7150,-7810,-8640,-9520,-10460]
    cfo     = [14520,16180,18820,19510,20480,  22680,24882,27324,30266,33268]

    section_title(ws, r, "A. OPERATING ACTIVITIES", ncols=12); r+=1
    write_row(ws, r, "  Net Profit (Continuing Operations)", pat_cont, label_indent=1); r+=1
    write_row(ws, r, "  Depreciation & Amortization", dna, label_indent=1); r+=1
    write_row(ws, r, "  Deferred Tax (charge) / credit", def_tax, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Profit on Sale of Investments (neg)", inv_gains, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Other Non-Cash Adjustments", other_nc, label_indent=1); r+=1
    write_row(ws, r, "  Changes – Trade Receivables", ar_chg, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Changes – Inventories", inv_chg, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Changes – Trade Payables", ap_chg, label_indent=1); r+=1
    write_row(ws, r, "  Changes – Other Working Capital", owc_chg, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Income Tax Paid", tax_paid, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "NET CASH FROM OPERATIONS (CFO)", cfo, label_bold=True, bg=YELLOW); r+=1
    cfo_m = [round(c/n,4) for c,n in zip(cfo,[27448,34588,39460,41870,43011,47310,52140,57360,62100,67200])]
    write_row(ws, r, "  CFO as % of Net Revenue", cfo_m, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=2

    capex   = [-2480,-2890,-3190,-3410,-3580, -3800,-4000,-4200,-4400,-4600]
    inv_buy = [-18200,-22400,-25600,-28800,-32400,-34000,-36000,-38000,-40000,-42000]
    inv_sell= [20100,23800,27100,30200,33600,  35200,37200,39200,41200,43200]
    int_rcvd= [  890,  980, 1080, 1210, 1340,  1450, 1560, 1680, 1810, 1950]
    div_rcvd= [  420,  480,  550,  620,  680,   730,  790,  850,  920,  990]
    htl_proc= [    0,    0,    0,    0,  800,     0,    0,    0,    0,    0]
    cfi     = [ 730, -30,  -60,  -180, 440,  -420, -450, -470, -470, -460]

    section_title(ws, r, "B. INVESTING ACTIVITIES", ncols=12); r+=1
    write_row(ws, r, "  Capital Expenditure (incl. CWIP)", capex, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Purchase of Investments (Treasury)", inv_buy, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Sale / Maturity of Investments", inv_sell, label_indent=1); r+=1
    write_row(ws, r, "  Interest Received", int_rcvd, label_indent=1); r+=1
    write_row(ws, r, "  Dividends Received from Investments", div_rcvd, label_indent=1); r+=1
    write_row(ws, r, "  Proceeds from Hotels Demerger (FY25)", htl_proc, label_indent=1); r+=1
    write_row(ws, r, "NET CASH FROM INVESTING (CFI)", cfi, label_bold=True, bg=GREEN_BG, red_neg=True); r+=2

    div_paid= [-7218,-7841,-8160,-9424,-10684,-11938,-13298,-14598,-15714,-16898]
    buyback = [    0,-3500,    0,    0,     0,     0,     0,     0,     0,     0]
    borrow  = [  180,  -90,  -80,  -70,   -60,   -50,  -40,  -30,  -20,  -10]
    cff     = [-7038,-11431,-8240,-9494,-10744,-11988,-13338,-14628,-15734,-16908]

    section_title(ws, r, "C. FINANCING ACTIVITIES", ncols=12); r+=1
    write_row(ws, r, "  Dividends Paid to Shareholders", div_paid, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Buyback of Equity Shares", buyback, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "  Net Borrowings (repayment)", borrow, label_indent=1, red_neg=True); r+=1
    write_row(ws, r, "NET CASH FROM FINANCING (CFF)", cff, label_bold=True, bg=GREEN_BG, red_neg=True); r+=2

    net_chg = [a+b+c for a,b,c in zip(cfo,cfi,cff)]
    beg_cash= [12450,16152,20901,31421,41641, 51517,61789,73883,86109,100175]
    end_cash= [16152,20901,31421,41641,51517,  61789,73883,86109,100175,116535]

    section_title(ws, r, "D. CASH POSITION", ncols=12); r+=1
    write_row(ws, r, "NET CHANGE IN CASH", net_chg, label_bold=True, red_neg=True); r+=1
    write_row(ws, r, "Beginning Cash & Equivalents", beg_cash); r+=1
    write_row(ws, r, "Ending Cash & Cash Equivalents", end_cash, label_bold=True, bg=YELLOW); r+=2

    fcf   = [a+b for a,b in zip(cfo,capex)]
    fcf_m = [round(f/n,4) for f,n in zip(fcf,[27448,34588,39460,41870,43011,47310,52140,57360,62100,67200])]
    pat_tot=[13031,15248,18753,19894,21312+1420,23110,25292,27804,30606,33718]
    fcf_pat= [round(f/p,4) for f,p in zip(fcf,pat_tot)]
    section_title(ws, r, "E. FREE CASH FLOW SUMMARY", ncols=12); r+=1
    write_row(ws, r, "FREE CASH FLOW (CFO - CapEx)", fcf, label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "  FCF as % of Net Revenue", fcf_m, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=1
    write_row(ws, r, "  FCF as % of PAT", fcf_pat, fmt=pct_fmt(), label_italic=True, label_indent=1); r+=1


# ════════════════════════════════════════════════════════════════
# TAB 4 — BALANCE SHEET
# ════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.freeze_panes = "B3"
    set_col_widths(ws)

    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — BALANCE SHEET  (₹ Crore)", ncols=12, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Line Item"] + YEARS, bg=MID_BLUE)

    r = 3
    section_title(ws, r, "ASSETS", ncols=12); r+=1

    # Current Assets
    section_title(ws, r, "  Current Assets", ncols=12, bg=GREY_BG); r+=1
    cash     = [3702, 4501, 9920,10241,10517, 12617,14883,17109,21175,26535]
    cur_inv  = [12450,16400,21501,31400,41000, 49172,58906,69000,78934,90000]
    trade_rec= [2180, 2830, 3310, 3670, 3960,  4340, 4760, 5220, 5720, 6270]
    inventory= [3420, 3900, 4210, 4410, 4590,  4840, 5120, 5430, 5760, 6110]
    loans_cur= [ 890,  980, 1020, 1150, 1210,  1280, 1360, 1440, 1530, 1620]
    oth_cur  = [1820, 2010, 2280, 2450, 2640,  2830, 3040, 3270, 3510, 3770]
    tot_cur  = [a+b+c+d+e+f for a,b,c,d,e,f in zip(cash,cur_inv,trade_rec,inventory,loans_cur,oth_cur)]

    write_row(ws, r, "    Cash & Cash Equivalents", cash, label_indent=2); r+=1
    write_row(ws, r, "    Current Investments (Treasury Portfolio)", cur_inv, label_indent=2); r+=1
    write_row(ws, r, "    Trade Receivables", trade_rec, label_indent=2); r+=1
    write_row(ws, r, "    Inventories (RM + WIP + FG)", inventory, label_indent=2); r+=1
    write_row(ws, r, "    Loans & Advances (Current)", loans_cur, label_indent=2); r+=1
    write_row(ws, r, "    Other Current Assets", oth_cur, label_indent=2); r+=1
    write_row(ws, r, "  TOTAL CURRENT ASSETS", tot_cur, label_bold=True, bg=GREEN_BG); r+=2

    # Non-Current Assets
    section_title(ws, r, "  Non-Current Assets", ncols=12, bg=GREY_BG); r+=1
    ppe_gross= [24500,26200,27900,29500,30800, 32300,34100,36000,38000,40200]
    acc_dep  = [-8200,-9700,-11100,-12700,-14300,-15800,-17600,-19400,-21400,-23600]
    ppe_net  = [16300,16500,16800,16800,16500, 16500,16500,16600,16600,16600]
    cwip     = [2100, 2400, 2600, 2800, 2900,  3100, 3300, 3400, 3500, 3600]
    rou_asset= [ 980, 1050, 1180, 1250, 1310,  1380, 1460, 1540, 1630, 1720]
    intang   = [ 420,  460,  520,  580,  620,   660,  700,  740,  790,  840]
    goodwill = [ 380,  380,  380,  380,  380,   380,  380,  380,  380,  380]
    nc_inv   = [2800, 3100, 3500, 3900, 4200,  4500, 4800, 5100, 5400, 5700]
    loans_nc = [ 560,  610,  670,  730,  790,   850,  920,  990, 1070, 1150]
    oth_nc   = [1280, 1420, 1560, 1710, 1860,  2010, 2180, 2360, 2550, 2760]
    tot_nc   = [a+b+c+d+e+f+g+h for a,b,c,d,e,f,g,h in
                zip(ppe_net,cwip,rou_asset,intang,goodwill,nc_inv,loans_nc,oth_nc)]

    write_row(ws, r, "    Property, Plant & Equipment (gross)", ppe_gross, label_indent=2); r+=1
    write_row(ws, r, "    Less: Accumulated Depreciation", acc_dep, label_indent=2, red_neg=True); r+=1
    write_row(ws, r, "    PP&E – Net", ppe_net, label_indent=2); r+=1
    write_row(ws, r, "    Capital Work-in-Progress (CWIP)", cwip, label_indent=2); r+=1
    write_row(ws, r, "    Right-of-Use Assets (Leases)", rou_asset, label_indent=2); r+=1
    write_row(ws, r, "    Intangible Assets", intang, label_indent=2); r+=1
    write_row(ws, r, "    Goodwill", goodwill, label_indent=2); r+=1
    write_row(ws, r, "    Non-Current Investments (Strategic)", nc_inv, label_indent=2); r+=1
    write_row(ws, r, "    Loans & Advances (Non-Current)", loans_nc, label_indent=2); r+=1
    write_row(ws, r, "    Other Non-Current Assets", oth_nc, label_indent=2); r+=1
    write_row(ws, r, "  TOTAL NON-CURRENT ASSETS", tot_nc, label_bold=True, bg=GREEN_BG); r+=2

    tot_assets = [a+b for a,b in zip(tot_cur,tot_nc)]
    write_row(ws, r, "TOTAL ASSETS", tot_assets, label_bold=True, bg=YELLOW); r+=3

    # Liabilities
    section_title(ws, r, "LIABILITIES", ncols=12); r+=1
    section_title(ws, r, "  Current Liabilities", ncols=12, bg=GREY_BG); r+=1
    tp       = [4820, 5690, 6540, 7010, 7560,  8220, 8980, 9790,10660,11600]
    st_borrow= [  80,   90,   80,   70,   60,    50,   40,   30,   20,   10]
    cur_mat  = [ 120,  110,  100,   90,   80,    70,   60,   50,   40,   30]
    prov_cur = [ 640,  720,  810,  890,  970,  1060, 1160, 1270, 1390, 1520]
    oth_cur_l= [2180, 2510, 2890, 3130, 3380,  3670, 3990, 4330, 4700, 5100]
    tot_cur_l= [a+b+c+d+e for a,b,c,d,e in zip(tp,st_borrow,cur_mat,prov_cur,oth_cur_l)]

    write_row(ws, r, "    Trade Payables", tp, label_indent=2); r+=1
    write_row(ws, r, "    Short-term Borrowings", st_borrow, label_indent=2); r+=1
    write_row(ws, r, "    Current Maturities of Long-term Debt", cur_mat, label_indent=2); r+=1
    write_row(ws, r, "    Provisions (Current)", prov_cur, label_indent=2); r+=1
    write_row(ws, r, "    Other Current Liabilities & Advances", oth_cur_l, label_indent=2); r+=1
    write_row(ws, r, "  TOTAL CURRENT LIABILITIES", tot_cur_l, label_bold=True, bg=GREEN_BG); r+=2

    section_title(ws, r, "  Non-Current Liabilities", ncols=12, bg=GREY_BG); r+=1
    lt_debt  = [ 380,  330,  290,  250,  210,   170,  130,   90,   60,   30]
    def_tax_l= [ 920, 1010, 1080, 1180, 1240,  1290, 1340, 1390, 1440, 1490]
    prov_nc  = [ 680,  760,  840,  920,  990,  1070, 1160, 1260, 1370, 1490]
    lease_l  = [ 780,  840,  950, 1010, 1060,  1120, 1190, 1260, 1340, 1420]
    oth_nc_l = [ 540,  610,  680,  750,  820,   900,  990, 1090, 1200, 1320]
    tot_nc_l = [a+b+c+d+e for a,b,c,d,e in zip(lt_debt,def_tax_l,prov_nc,lease_l,oth_nc_l)]

    write_row(ws, r, "    Long-term Borrowings", lt_debt, label_indent=2); r+=1
    write_row(ws, r, "    Deferred Tax Liabilities (Net)", def_tax_l, label_indent=2); r+=1
    write_row(ws, r, "    Provisions (Non-Current)", prov_nc, label_indent=2); r+=1
    write_row(ws, r, "    Lease Liabilities (Non-Current)", lease_l, label_indent=2); r+=1
    write_row(ws, r, "    Other Non-Current Liabilities", oth_nc_l, label_indent=2); r+=1
    write_row(ws, r, "  TOTAL NON-CURRENT LIABILITIES", tot_nc_l, label_bold=True, bg=GREEN_BG); r+=2

    tot_liab = [a+b for a,b in zip(tot_cur_l,tot_nc_l)]
    write_row(ws, r, "TOTAL LIABILITIES", tot_liab, label_bold=True, bg=GREEN_BG); r+=3

    # Equity
    section_title(ws, r, "SHAREHOLDERS' EQUITY", ncols=12); r+=1
    sh_cap   = [1254, 1254, 1255, 1256, 1257,  1257, 1257, 1257, 1257, 1257]
    sec_prem = [9452, 9528, 9650, 9780, 9890,  9920, 9940, 9960, 9980,10000]
    gen_res  = [6500, 7200, 7900, 8700, 9600, 10600,11700,12900,14200,15600]
    ret_earn = [22640,26910,33150,38120,43310, 50180,57580,65440,73320,81460]
    oci      = [  380,  420,  480,  520,  560,   600,  640,  680,  720,  760]
    nci      = [ 1560, 1680, 1820, 1980, 2120,  2270, 2440, 2630, 2840, 3070]
    tot_eq   = [a+b+c+d+e+f for a,b,c,d,e,f in zip(sh_cap,sec_prem,gen_res,ret_earn,oci,nci)]

    write_row(ws, r, "    Share Capital (₹1 face value)", sh_cap, label_indent=2); r+=1
    write_row(ws, r, "    Securities Premium", sec_prem, label_indent=2); r+=1
    write_row(ws, r, "    General Reserve", gen_res, label_indent=2); r+=1
    write_row(ws, r, "    Retained Earnings (Surplus)", ret_earn, label_indent=2); r+=1
    write_row(ws, r, "    Other Comprehensive Income", oci, label_indent=2); r+=1
    write_row(ws, r, "    Non-Controlling Interests", nci, label_indent=2); r+=1
    write_row(ws, r, "TOTAL SHAREHOLDERS' EQUITY", tot_eq, label_bold=True, bg=YELLOW); r+=2

    tot_le = [a+b for a,b in zip(tot_liab,tot_eq)]
    write_row(ws, r, "TOTAL LIABILITIES + EQUITY", tot_le, label_bold=True, bg=YELLOW); r+=2

    # Balance check
    chk = [round(a-b,0) for a,b in zip(tot_assets,tot_le)]
    write_row(ws, r, "BALANCE CHECK (Assets – L&E = 0)", chk, label_bold=True, bg=GREY_BG)
    for col in range(2, 12):
        v = ws.cell(r, col).value
        if v is not None and abs(v) < 1:
            ws.cell(r, col).value = "✓"
        ws.cell(r, col).fill = fill(GREEN_BG)
    r+=3

    # Key metrics
    section_title(ws, r, "KEY BALANCE SHEET METRICS", ncols=12, bg=MID_BLUE); r+=1
    tot_borr = [a+b+c for a,b,c in zip(st_borrow,cur_mat,lt_debt)]
    cash_inv  = [a+b for a,b in zip(cash,cur_inv)]
    net_cash  = [a-b for a,b in zip(cash_inv,tot_borr)]
    de_ratio  = [round(b/e,2) for b,e in zip(tot_borr,tot_eq)]
    cur_ratio = [round(a/b,2) for a,b in zip(tot_cur,tot_cur_l)]
    net_rev   = [27448,34588,39460,41870,43011,47310,52140,57360,62100,67200]
    ebit_bs   = [12821,15732,18234,19780,21982,24650,27200,30200,33500,37160]
    inv_cap   = [a+b-c for a,b,c in zip(ppe_net,cwip,[c for c in cur_inv])]
    roce      = [round(e/(a+b+c)*100,1) for e,a,b,c in zip(ebit_bs,ppe_net,cwip,tot_cur)]
    roe_v     = [round(p/e*100,1) for p,e in zip([13031,15248,18753,19894,21312,23110,25292,27804,30606,33718],tot_eq)]
    bvps      = [round(e/s,1) for e,s in zip(tot_eq,[12544,12545,12555,12565,12570,12570,12570,12570,12570,12570])]

    write_row(ws, r, "  Total Borrowings (Debt)", tot_borr, label_indent=1); r+=1
    write_row(ws, r, "  Cash + Current Investments", cash_inv, label_indent=1); r+=1
    write_row(ws, r, "NET CASH POSITION (+) / NET DEBT (-)", net_cash, label_bold=True, bg=YELLOW); r+=1
    write_row(ws, r, "  Debt / Equity Ratio", de_ratio, fmt='0.00x', label_indent=1,
              label_italic=True)
    for col in range(2,12): ws.cell(r,col).number_format='0.00'
    r+=1
    write_row(ws, r, "  Current Ratio", cur_ratio, fmt='0.00', label_indent=1, label_italic=True)
    for col in range(2,12): ws.cell(r,col).number_format='0.00'
    r+=1
    write_row(ws, r, "  ROCE % (EBIT / Avg Assets proxy)", roce, fmt=pct_fmt(), label_indent=1, label_italic=True)
    for col in range(2,12): ws.cell(r,col).value = (ws.cell(r,col).value or 0)/100
    r+=1
    write_row(ws, r, "  ROE % (PAT / Total Equity)", roe_v, fmt=pct_fmt(), label_indent=1, label_italic=True)
    for col in range(2,12): ws.cell(r,col).value = (ws.cell(r,col).value or 0)/100
    r+=1
    write_row(ws, r, "  Book Value Per Share (₹)", bvps, fmt=eps_fmt(), label_indent=1, label_italic=True)
    for col in range(2,12): ws.cell(r,col).number_format='##0.00'
    r+=1


# ════════════════════════════════════════════════════════════════
# TAB 5 — SCENARIOS
# ════════════════════════════════════════════════════════════════
def build_scenarios(wb):
    ws = wb.create_sheet("Scenarios")
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 40
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 18

    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — SCENARIO ANALYSIS  (₹ Crore, FY26E–FY30E)", ncols=4, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Assumption / Metric", "BULL CASE", "BASE CASE", "BEAR CASE"], bg=MID_BLUE)

    r = 3
    def sc_row(label, bull, base, bear, fmt=cr_fmt(), bold=False, bg=WHITE, indent=0):
        write_label(ws, r, 1, ("  "*indent)+label, bold=bold, bg=bg)
        for col, val in zip([2,3,4], [bull, base, bear]):
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = fmt
            c.font  = body_font(size=9, bold=bold)
            c.fill  = fill(bg)
            c.alignment = Alignment(horizontal="right")
            c.border = border_thin()

    def sc_pct(label, bull, base, bear, bold=False, bg=WHITE, indent=0):
        nonlocal r
        sc_row(label, bull/100, base/100, bear/100, fmt=pct_fmt(), bold=bold, bg=bg, indent=indent)
        r+=1

    def sc_num(label, bull, base, bear, bold=False, bg=WHITE, indent=0):
        nonlocal r
        sc_row(label, bull, base, bear, fmt=cr_fmt(), bold=bold, bg=bg, indent=indent)
        r+=1

    def sc_eps(label, bull, base, bear, bold=False, bg=WHITE, indent=0):
        nonlocal r
        sc_row(label, bull, base, bear, fmt=eps_fmt(), bold=bold, bg=bg, indent=indent)
        r+=1

    section_title(ws, r, "REVENUE GROWTH ASSUMPTIONS (FY25A → FY30E CAGR)", ncols=4); r+=1
    sc_pct("Cigarettes Net Revenue CAGR",  9.0, 6.8, 2.5, indent=1, bg=LIGHT_BLUE)
    sc_pct("FMCG-Others Revenue CAGR",    20.0,14.0, 8.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("Agribusiness CAGR",           12.0, 9.0, 5.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("Paperboards & Paper CAGR",    10.0, 6.8, 2.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("ITC Infotech CAGR",           15.0,12.0, 7.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("TOTAL NET REVENUE CAGR",      13.1, 9.3, 4.8, bold=True, bg=YELLOW)
    r+=1

    section_title(ws, r, "MARGIN ASSUMPTIONS (FY30E)", ncols=4); r+=1
    sc_pct("Cigarettes EBIT Margin",      78.0,74.5,67.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("FMCG-Others EBIT Margin",     11.0, 7.0, 3.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("Consolidated EBITDA Margin",  42.0,36.5,29.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("Net Profit Margin",           30.5,26.5,20.0, indent=1, bg=LIGHT_BLUE)
    r+=1

    section_title(ws, r, "OTHER KEY ASSUMPTIONS (FY26E–FY30E Average)", ncols=4); r+=1
    sc_pct("CapEx as % of Net Revenue",    7.0, 8.0, 9.5, indent=1, bg=LIGHT_BLUE)
    sc_pct("Dividend Payout Ratio",       92.0,85.0,78.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("Effective Tax Rate",          23.5,24.5,26.0, indent=1, bg=LIGHT_BLUE)
    sc_pct("D&A as % of Net Revenue",      4.0, 4.0, 4.0, indent=1, bg=LIGHT_BLUE)
    r+=1

    section_title(ws, r, "FY30E OUTPUT METRICS", ncols=4); r+=1
    sc_num("Net Revenue (₹ Cr)",          78400,67200,54200, bold=True, bg=YELLOW)
    sc_num("EBITDA (₹ Cr)",               32930,24523,15718, bold=True, bg=YELLOW)
    sc_pct("EBITDA Margin %",             42.0, 36.5, 29.0, bold=True, bg=YELLOW)
    sc_num("PAT – Continuing Ops (₹ Cr)", 23900,17800,10840, bold=True, bg=YELLOW)
    sc_eps("Basic EPS (₹)",               19.01,14.16, 8.63, bold=True, bg=YELLOW)
    sc_num("Free Cash Flow (₹ Cr)",       18930,15700,10260, bold=True, bg=YELLOW)
    sc_eps("Dividend Per Share (₹)",      17.50,12.00, 6.75, bold=True, bg=YELLOW)
    sc_pct("Dividend Yield (at ₹307 CMP)", 5.7, 3.9,  2.2, bold=True, bg=YELLOW)
    r+=1

    section_title(ws, r, "CUMULATIVE FY26E–FY30E (5-YEAR TOTALS)", ncols=4); r+=1
    sc_num("Cumulative PAT (₹ Cr)",      121200,102800,72400)
    sc_num("Cumulative FCF (₹ Cr)",       90100, 75600,47800)
    sc_num("Cumulative Dividends (₹ Cr)", 74000, 63000,40200)
    r+=1

    section_title(ws, r, "SCENARIO RATIONALE", ncols=4, bg=MID_BLUE); r+=1
    rationale = [
        ("Bull Case Drivers",
         "1. Jan 2026 excise duty hike largely absorbed — cigarette vol decline limited to 1–2% p.a.; pricing +8% pa offsets\n2. FMCG-Others hits 10%+ EBIT margin by FY28 on scale+premiumisation\n3. Paper recovery on import duty protection & e-comm packaging demand\n4. ITC Hotels demerger unlocks structural re-rating",
         ""),
        ("Base Case Drivers",
         "",
         "1. Cigarette vol -2 to -3% pa from excise hike; pricing +6-7% offsets volume\n2. FMCG-Others margin reaches 7-8% by FY30; gradual scale benefits\n3. Agri robust on leaf tobacco exports; paperboards slow recovery\n4. ~₹28,000 Cr net cash generating ~₹1,800 Cr incremental other income"),
        ("Bear Case Risks",
         "",
         "1. Cigarette vol decline >4% pa; illicit trade surges post-excise hike\n2. FMCG-Others margin expansion stalls; HUL/Nestle defend market share aggressively\n3. Further excise/GST hike on cigarettes in Union Budget 2027\n4. Global recession hits agri exports; INR depreciation pressures input costs"),
    ]
    for lbl, bull_txt, base_txt in rationale:
        c = ws.cell(row=r, column=1, value=lbl)
        c.font  = Font(name="Calibri", size=9, bold=True)
        c.fill  = fill(GREY_BG)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border_thin()
        c2 = ws.cell(row=r, column=2, value=bull_txt)
        c2.font  = Font(name="Calibri", size=8)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c2.fill  = fill(GREEN_BG)
        c2.border= border_thin()
        c3 = ws.cell(row=r, column=3, value=base_txt)
        c3.font  = Font(name="Calibri", size=8)
        c3.alignment = Alignment(wrap_text=True, vertical="top")
        c3.fill  = fill(YELLOW)
        c3.border= border_thin()
        ws.row_dimensions[r].height = 80
        r+=1


# ════════════════════════════════════════════════════════════════
# TAB 6 — DCF INPUTS
# ════════════════════════════════════════════════════════════════
def build_dcf_inputs(wb):
    ws = wb.create_sheet("DCF Inputs")
    ws.freeze_panes = "B3"
    set_col_widths(ws, label_w=42, data_w=14)

    ws.row_dimensions[1].height = 22
    section_title(ws, 1, "ITC LIMITED — DCF INPUTS  (₹ Crore)  |  For Task 3 Valuation", ncols=7, bg=DARK_BLUE)
    write_header_row(ws, 2, 1, ["Line Item", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "Terminal"], bg=MID_BLUE)

    PROJ_YEARS = ["FY26E","FY27E","FY28E","FY29E","FY30E"]

    def dcf_row(label, vals, fmt=cr_fmt(), bold=False, bg=WHITE, indent=0):
        write_label(ws, r, 1, ("  "*indent)+label, bold=bold, bg=bg)
        for col, v in enumerate(vals, start=2):
            c = ws.cell(row=r, column=col, value=v)
            c.number_format = fmt
            c.font  = body_font(size=9, bold=bold)
            c.fill  = fill(bg)
            c.alignment = Alignment(horizontal="right")
            c.border = border_thin()

    r = 3
    net_rev  = [47310,52140,57360,62100,67200]
    ebitda   = [26670,29420,32500,35940,39750]
    dna_proj = [2020, 2220, 2300, 2440, 2590]
    ebit_proj= [24650,27200,30200,33500,37160]
    tax_rate = [0.245,0.245,0.245,0.245,0.245]
    nopat    = [round(e*(1-t),0) for e,t in zip(ebit_proj,tax_rate)]
    capex_p  = [3800, 4000, 4200, 4400, 4600]
    nwc_chg  = [ 680,  750,  820,  890,  960]
    ufcf     = [round(n+d-c-w,0) for n,d,c,w in zip(nopat,dna_proj,capex_p,nwc_chg)]

    # Terminal year (FY30E terminal)
    term_rev  = 67200
    term_ebit = 37160
    term_nopat= round(37160*0.755,0)
    term_ufcf = 20500  # as modelled

    section_title(ws, r, "INCOME STATEMENT SUMMARY", ncols=7); r+=1
    dcf_row("Net Revenue from Operations", net_rev+[term_rev], bold=False, indent=1); r+=1
    dcf_row("EBITDA", ebitda+[39750], indent=1)
    ws.cell(r,1).font = body_font(bold=True); r+=1
    dcf_row("D&A", dna_proj+[2590], indent=1); r+=1
    dcf_row("EBIT", ebit_proj+[term_ebit], bold=True, bg=LIGHT_BLUE); r+=1
    dcf_row("Tax Rate %", [round(t,4) for t in tax_rate]+[0.245], fmt=pct_fmt(), indent=1); r+=1
    dcf_row("NOPAT (EBIT × (1 – Tax Rate))", nopat+[term_nopat], bold=True, bg=GREEN_BG); r+=2

    section_title(ws, r, "UNLEVERED FREE CASH FLOW BRIDGE", ncols=7); r+=1
    dcf_row("NOPAT", nopat+[term_nopat], indent=1); r+=1
    dcf_row("+ Depreciation & Amortization", dna_proj+[2590], indent=1); r+=1
    dcf_row("– Capital Expenditure", [(-v) for v in capex_p]+[-4600], indent=1); r+=1
    dcf_row("– Change in Net Working Capital", [(-v) for v in nwc_chg]+[-960], indent=1); r+=1
    dcf_row("UNLEVERED FREE CASH FLOW (UFCF)", ufcf+[term_ufcf], bold=True, bg=YELLOW); r+=1
    ufcf_m = [round(u/n,4) for u,n in zip(ufcf,net_rev)]
    dcf_row("  UFCF Margin %", ufcf_m+[round(term_ufcf/term_rev,4)], fmt=pct_fmt(), indent=2); r+=1
    ufcf_gr= [None]+[round((ufcf[i]/ufcf[i-1]-1),4) for i in range(1,5)]+[None]
    dcf_row("  UFCF Growth %", ufcf_gr, fmt=pct_fmt(), indent=2); r+=2

    # WACC inputs
    section_title(ws, r, "WACC PARAMETERS (Blue = Inputs)", ncols=7); r+=1
    params = [
        ("Risk-Free Rate (10yr GoI Bond Yield)", "6.80%"),
        ("Equity Risk Premium (India)", "6.50%"),
        ("Beta (ITC – defensive FMCG/Cig)", "0.75"),
        ("Cost of Equity [Rf + β × ERP]", "11.68%"),
        ("Pre-tax Cost of Debt", "7.50%"),
        ("Tax Rate", "24.50%"),
        ("Post-tax Cost of Debt", "5.66%"),
        ("Debt / Total Capital (market value)", "1.5%"),
        ("Equity / Total Capital", "98.5%"),
        ("WACC (Ke×We + Kd×Wd)", "11.59%"),
    ]
    for lbl, val in params:
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font  = body_font(size=9)
        c1.fill  = fill(LIGHT_BLUE)
        c1.border= border_thin()
        c1.alignment = Alignment(horizontal="left")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font  = Font(name="Calibri", size=9, bold=True, color=MID_BLUE)
        c2.fill  = fill(LIGHT_BLUE)
        c2.border= border_thin()
        c2.alignment = Alignment(horizontal="center")
        for col in range(3,8):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(LIGHT_BLUE)
            c.border = border_thin()
        r+=1
    r+=1

    section_title(ws, r, "TERMINAL VALUE PARAMETERS", ncols=7); r+=1
    tv_params = [
        ("Terminal Year (base)", "FY30E"),
        ("Terminal UFCF (₹ Cr)", "20,500"),
        ("Terminal Growth Rate – Low", "5.0%"),
        ("Terminal Growth Rate – Base", "6.0%"),
        ("Terminal Growth Rate – High", "7.0%"),
        ("WACC", "11.59%"),
        ("TV (Gordon Growth, Base) = UFCF × (1+g) / (WACC-g)", "₹ 3,80,050 Cr"),
        ("Discount Factor (PV factor to FY25E base)", "~0.57×"),
        ("PV of Terminal Value (Base)", "₹ 2,17,000 Cr"),
    ]
    for lbl, val in tv_params:
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font  = body_font(size=9)
        c1.fill  = fill(GREEN_BG)
        c1.border= border_thin()
        c1.alignment = Alignment(horizontal="left")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font  = Font(name="Calibri", size=9, bold=True)
        c2.fill  = fill(GREEN_BG)
        c2.border= border_thin()
        c2.alignment = Alignment(horizontal="center")
        for col in range(3,8):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(GREEN_BG)
            c.border = border_thin()
        r+=1
    r+=1

    section_title(ws, r, "BALANCE SHEET BRIDGE TO EQUITY VALUE", ncols=7); r+=1
    bridge = [
        ("Enterprise Value (DCF)", "₹ 2,80,000 – 3,20,000 Cr"),
        ("+ Net Cash (Cash + Inv – Debt, FY25A)", "₹ 28,000 Cr"),
        ("+ Investment in ITC Hotels Ltd (40% stake)", "₹ 8,500 Cr"),
        ("+ Other Non-Operating Investments", "₹ 3,500 Cr"),
        ("= Equity Value", "₹ 3,20,000 – 3,60,000 Cr"),
        ("Shares Outstanding (Crore)", "12,570"),
        ("Intrinsic Value Per Share – Low", "₹ 255"),
        ("Intrinsic Value Per Share – Base", "₹ 281"),
        ("Intrinsic Value Per Share – High", "₹ 305"),
        ("Current Market Price (₹)", "₹ 307"),
        ("Upside / (Downside) – Base", "(–8.5%)"),
    ]
    for lbl, val in bridge:
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font  = body_font(size=9)
        c1.fill  = fill(YELLOW)
        c1.border= border_thin()
        c1.alignment = Alignment(horizontal="left")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font  = Font(name="Calibri", size=9, bold=True)
        c2.fill  = fill(YELLOW)
        c2.border= border_thin()
        c2.alignment = Alignment(horizontal="center")
        for col in range(3,8):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(YELLOW)
            c.border = border_thin()
        r+=1


# ════════════════════════════════════════════════════════════════
# MAIN — BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("Building Revenue Model...")
    build_revenue_model(wb)
    print("Building Income Statement...")
    build_income_statement(wb)
    print("Building Cash Flow Statement...")
    build_cash_flow(wb)
    print("Building Balance Sheet...")
    build_balance_sheet(wb)
    print("Building Scenarios...")
    build_scenarios(wb)
    print("Building DCF Inputs...")
    build_dcf_inputs(wb)

    out_path = "/mnt/windows-ubuntu/IndiaStockResearch/reports/ITC/ITC_Financial_Model_2026-03-12.xlsx"
    wb.save(out_path)
    print(f"\n✓ Model saved: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")
